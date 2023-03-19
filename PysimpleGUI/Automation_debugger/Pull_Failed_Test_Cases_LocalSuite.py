import time
from pyvirtualdisplay import Display
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook


class Pull_Failed_Test_Cases_LocalSuite:

    def __init__(self):
        display = Display(visible=False, size=(800, 800))
        display.start()
        self.chromeOptions = Options()
        self.chromeOptions.add_argument("--no-sandbox")
        self.chromeOptions.add_argument("--disable-extensions")
        self.chromeOptions.add_argument("--headless")
        self.chromeOptions.add_experimental_option("excludeSwitches", ["enable-logging"])
        self.chromeOptions.binary_location = "/usr/bin/google-chrome"
        self.chromeOptions.add_argument("--remote-debugging-port=9222")
        self.chromeOptions.add_argument("--disable-dev-shm-usage")
        self.wb = Workbook()
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    def launch_browser(self, html_path, excel_path):
        print(f"htm file path is file:// + {html_path}")
        self.driver.get("file://" + html_path)
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(30)
        result = self.read_data(excel_path)
        return result

    def read_data(self, excel_path):
        try:
            ws = self.wb.active
            titles = {'A': 'Parent Suite', 'B': 'Suite', 'C': 'Test Case', 'D': 'Failure Reason'}
            ws.append(titles)
            count = 2
            # Total number of suites
            suite_list = self.driver.find_element(By.XPATH, "//span[contains(text(), "
                                                             "\"SUITE\")]//parent::div//following-sibling::span[2][@class=\"label "
                                                             "fail\"]//following-sibling::span[@class=\"name\"]")
            print(suite_list)
            print(f"Suite list in the start is {suite_list}")

            print(f"Suite list actual is {suite_list.text}")

            type_of_column = "suites"

            underscore_suite = ""

            suite_list1 = self.driver.find_element(By.XPATH,
                                                    "//span[contains(text(), \"SUITE\")]//parent::div//following-sibling::span[2][@class=\"label fail\"]//following-sibling::span[@class=\"name\"]")
            s2 = suite_list1.text.strip()
            id_type = self.driver.find_element(By.XPATH,
                                               "//span[text()='" + suite_list1.text + "']//parent::div//parent::div//parent::div[@class=\"suite\"]").get_attribute(
                "id")
            print(f"id_type is {id_type}")
            dashCount = id_type.count("-")
            print(f"Value of dashcount is {dashCount}")
            normal_suite = s2
            print(f"Normal suite value is {normal_suite}")
            ws.cell(row=count, column=2).value = normal_suite
            # Test cases
            myXpath = "//span[@class=\"name\" and " \
                      "text()='" + normal_suite + "']//parent::div//parent::div//parent::div//span[text(" \
                                                  ")=\"TEST\"]//parent::div//following-sibling::span[2][" \
                                                  "@class=\"label fail\"]//following-sibling::span[" \
                                                  "@class=\"name\"]"
            print(f"My xpath is {myXpath}")
            print(f"Test suite count is {count}")
            test_case_list = self.driver.find_elements(By.XPATH, myXpath)
            toggle_xpath = "((//span[@class=\"name\" and " \
                           "text()='" + normal_suite + "']//parent::div//parent::div//parent::div//div[" \
                                                       "@class=\"element-header-toggle\"])[1]"
            try:
                if len(test_case_list) == 0 and len(underscore_suite) != 0 and len(normal_suite) != 0:
                    self.driver.find_element(By.XPATH, toggle_xpath).click()
                    test_case_list = self.driver.find_elements(By.XPATH, myXpath)
            except Exception as e:
                print("Unable to click on child suite toggle")
                print(e)
            print(f"Sending test case list is {test_case_list}")
            count = self.edit_excel_sheet(test_case_list, ws, count + 1, type_of_column, underscore_suite,
                                          normal_suite)
            count = count + 1
            self.wb.save(excel_path + ".xlsx")
            return True
        except Exception as e:
            print(e)
            return False

    def edit_excel_sheet(self, test_case_list, ws, count, type_of_column, underscore_suite, normal_suite):
        try:
            for c in test_case_list:
                words1 = [c.text]
                print(f"Output of words1 is {words1}")
                ws.cell(row=count, column=3).value = c.text
                myXpath2 = "((//span[@class=\"name\" and text(" \
                           ")='" + normal_suite + "']//parent::div//parent::div//parent::div//span[text()=" + "\"" + \
                           c.text + "\"" + "]//parent::div//parent::div//parent::div//div[@class=\"children " \
                                           "populated\"])[1]//tr/td[@class=\"message\"])[1]"
                testcase_toggle_xpath = "(//span[" \
                                        "@class=\"name\" and text()='" + normal_suite + \
                                        "']//parent::div//parent::div//parent::div//span[text()=" + "\"" + c.text + "\"" + "]//parent::div//parent::div//parent::div//div[@class=\"element-header-toggle\"])[1] "

                toggle_state_doc = self.driver.find_element(By.XPATH,
                                                            "(//span[text()=" + "\"" + c.text + "\"" + "]//parent::div//parent::div//parent::div//div[@class=\"element-header-toggle\"])[1]//parent::div").get_attribute(
                    "class")
                print(f"Toggle_state_doc is {toggle_state_doc}")
                if toggle_state_doc == "element-header closed":
                    self.driver.find_element(By.XPATH, testcase_toggle_xpath).click()
                Failed_message = self.driver.find_element(By.XPATH, myXpath2)
                print("Failed test case info is " + Failed_message.text)
                current_tc_data = ws.cell(row=count, column=3)
                ws.delete_rows(count)
                append_dict = {'C': current_tc_data.value, 'D': Failed_message.text}
                ws.append(append_dict)
                count = count + 1
            return count
        except Exception as e:
            print(e)
            print("Looks like" + type_of_column + " is/are not available")
            return count

    def close_browser(self):
        self.driver.close()
