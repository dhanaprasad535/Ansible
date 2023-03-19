import random

import PySimpleGUI as sg
from openpyxl import load_workbook

import Pull_Failed_Test_Cases
import Pull_Failed_Test_Cases_Local
import Pull_Failed_Test_Cases_LocalSuite
import Read_From_Excel


def create_window():
    sg.theme('SystemDefault')
    font = ("Lora", 11, 'italic')
    sg.set_options(font=font)

    label1 = sg.Text("Select HTML log file location:", text_color='black')
    input1 = sg.Input(key="html")
    choose1 = sg.FileBrowse("Choose", key="choose1")
    label2 = sg.Text("Name of the excel sheet(example:test):", text_color='black')
    input2 = sg.Input(key="excel")
    box = sg.Multiline(enable_events=True, size=(80, 15), key="summary")
    filetype_button = sg.Spin(["Jenkins", "Local-Folder", "Local-Suite"], size=(13, 1), key='-FILETYPE-')
    help_button = sg.Button("Help", key="help")
    submit = sg.Button("Submit", key="submit")
    copy = sg.Button("Create a text file", key="copy")
    progress = sg.Text("", key="-PROGRESS-", justification='center', expand_x=True, enable_events=True)
    output = sg.Text(key="output")

    return sg.Window("Failed Cases Report Generator",
                     layout=[[label1, input1, choose1], [label2, input2], [box],
                             [help_button, submit, progress, sg.Push(), filetype_button, copy], [output]],
                     finalize=True)


window = create_window()

while True:
    event, values = window.read()
    print(event, values)
    window["-PROGRESS-"].update(value="In Progress") # This part is not working
    if event == "submit":
        if values["excel"] != '' and values["html"] != '':
            obj = ''
            if values["-FILETYPE-"] == "Jenkins":
                obj = Pull_Failed_Test_Cases.Pull_Failed_Test_Cases()
            elif values["-FILETYPE-"] == "Local-Folder":
                obj = Pull_Failed_Test_Cases_Local.Pull_Failed_Test_Cases_Local()
            elif values["-FILETYPE-"] == "Local-Suite":
                obj = Pull_Failed_Test_Cases_LocalSuite.Pull_Failed_Test_Cases_LocalSuite()
            result = obj.launch_browser(values["html"], values["excel"])
            if result:
                window["output"].update(value="Data has been fetched successfully")
                excel = values["excel"]
                wb = load_workbook(excel + ".xlsx")
                ws = wb.active

                print(f"Max row in the active sheet is {ws.max_row}")
                print(f"Max column in the active sheet is {ws.max_column}")

                testcase_failmsg_dict = {}
                for i in range(2, ws.max_row + 1):
                    message = ws.cell(row=i, column=4)
                    if message.value is not None:
                        print(
                            f"Test case {ws.cell(row=i, column=3).value} is failed with Failure message {message.value}")
                        tc = ws.cell(row=i, column=3).value + str(random.randrange(100, 999, 3))
                        testcase_failmsg_dict[tc] = message.value
                print(f" testcase_failmsg_dict is {testcase_failmsg_dict}")
                read_excel_dict = Read_From_Excel.read_from_excel(testcase_failmsg_dict)
                message = ''
                message_new = ''
                sum1 = 0
                for k, v in read_excel_dict.items():
                    failure_msg = list(k.strip("[").strip("]").strip("''").split(","))[-1]
                    a = list(k.strip("[").strip("]").strip("''").split(","))[:-1]
                    result = str(a)[1:-1]
                    result = result.replace("'", '')
                    last3_char = result[-4:-1]
                    last2_char = result[-3:-1]
                    if last3_char.isnumeric():
                        result = result.replace(last3_char, '')
                    if last2_char.isnumeric():
                        result = result.replace(last2_char, '')
                    message = '{} testcase(s) {} is/are failed with error {}'.format(v,
                                                                                     result.replace("''", '').replace(
                                                                                         ",", '').strip(" "),
                                                                                     failure_msg)
                    sum1 = sum1 + v
                    message_new = message_new + "\n\n" + message
                    window["summary"].update(value=message_new)
                message_new = message_new + "\n\n" + f"Total cases failed count is {sum1}"

                window["summary"].update(value=message_new)
                window["-PROGRESS-"].update(value="Completed")

            else:
                window["output"].update(value="Sorry. Unable to fetch the data")
                window["-PROGRESS-"].update(value="Completed")
            obj.close_browser()
        else:
            sg.popup("You should provide the HTML file and excel sheet name before clicking on submit button",
                     font=('Lora', 10), title="Error", text_color='black')
    elif event == "help":
        sg.popup("Select the automation log HTML file and the name of excel sheet that you want your failed test cases "
                 "info to be copied to and click on Submit\nAt the end, you get an excel sheet with all the failed test cases"
                 " and count of failed test cases will also be displayed in the text box", font=('Lora', 10),
                 title="Help Popup", text_color='black')
    elif event == 'copy':
        with open(values["excel"] + '.txt', 'w+') as file:
            savedText1 = file.write(values["summary"])
        file.close()

    elif event == sg.WINDOW_CLOSED:
        break

window.close()
